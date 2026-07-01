from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash
)

from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user
)

from werkzeug.security import generate_password_hash, check_password_hash

from config import Config
from models import db, Usuario, Gasto
from datetime import datetime
from sqlalchemy import or_
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from flask import send_file
from io import BytesIO
import os

app = Flask(__name__)
app.config.from_object(Config)

os.makedirs("database", exist_ok=True)
os.makedirs("backups", exist_ok=True)
os.makedirs("exports", exist_ok=True)
os.makedirs("uploads", exist_ok=True)

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))


@app.route("/")
def raiz():
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        usuario = request.form["usuario"]
        password = request.form["password"]

        user = Usuario.query.filter_by(
            usuario=usuario
        ).first()

        if user and check_password_hash(
                user.password,
                password):

            login_user(user)

            return redirect(
                url_for("inicio")
            )

        flash(
            "Usuario o contraseña incorrectos",
            "danger"
        )

    return render_template("login.html")


@app.route("/inicio")
@login_required
def inicio():

    return redirect(
        url_for("gastos")
)   

@app.route("/logout")
@login_required
def logout():

    logout_user()

    return redirect(
        url_for("login")
    )
@app.route("/gastos")
@login_required
def gastos():

    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    responsable = request.args.get("responsable")
    buscar = request.args.get("buscar")
    orden = request.args.get("orden", "fecha_desc")
    estado = request.args.get("estado", "todos")

    consulta = Gasto.query

    if fecha_inicio:
        consulta = consulta.filter(
            Gasto.fecha >= datetime.strptime(
                fecha_inicio,
                "%Y-%m-%d"
            ).date()
        )

    if fecha_fin:
        consulta = consulta.filter(
            Gasto.fecha <= datetime.strptime(
                fecha_fin,
                "%Y-%m-%d"
            ).date()
        )

    if responsable:
        consulta = consulta.filter(
            Gasto.responsable == responsable
        )

    if buscar:
        consulta = consulta.filter(
            or_(
                Gasto.pagado_a.contains(buscar),
                Gasto.concepto.contains(buscar),
                Gasto.observaciones.contains(buscar)
            )
        )

    # -----------------------------
    # FILTRO POR ESTADO
    # -----------------------------

    if estado == "pendientes":
        consulta = consulta.filter(
            Gasto.comprobante == False
        )

    elif estado == "elaborados":
        consulta = consulta.filter(
            Gasto.comprobante == True
        )

    # -----------------------------
    # ORDENAMIENTO
    # -----------------------------

    if orden == "fecha_asc":

        consulta = consulta.order_by(
            Gasto.fecha.asc()
        )

    elif orden == "fecha_desc":

        consulta = consulta.order_by(
            Gasto.fecha.desc()
        )

    elif orden == "importe_asc":

        consulta = consulta.order_by(
            Gasto.importe.asc()
        )

    elif orden == "importe_desc":

        consulta = consulta.order_by(
            Gasto.importe.desc()
        )

    elif orden == "responsable_asc":

        consulta = consulta.order_by(
            Gasto.responsable.asc()
        )

    elif orden == "responsable_desc":

        consulta = consulta.order_by(
            Gasto.responsable.desc()
        )

    else:

        consulta = consulta.order_by(
            Gasto.id.desc()
        )

    gastos = consulta.all()

    total = sum(
        g.importe for g in gastos
    )

    total_registros = len(gastos)

    elaborados = Gasto.query.filter_by(
        comprobante=True
    ).count()

    pendientes = Gasto.query.filter_by(
        comprobante=False
    ).count()

    return render_template(

        "index.html",

        gastos=gastos,

        total=total,

        total_registros=total_registros,

        elaborados=elaborados,

        pendientes=pendientes,

        fecha_inicio=fecha_inicio,

        fecha_fin=fecha_fin,

        responsable=responsable,

        buscar=buscar,

        orden=orden,

        estado=estado

    )
@app.route("/nuevo", methods=["GET", "POST"])
@login_required
def nuevo():

    if request.method == "POST":

        gasto = Gasto(
            fecha=datetime.strptime(
                request.form["fecha"],
                "%Y-%m-%d"
            ),

            pagado_a=request.form["pagado_a"],

            concepto=request.form["concepto"],

            observaciones=request.form["observaciones"],

            responsable=request.form["responsable"],

            importe=float(
                request.form["importe"]
            ),

            usuario_id=current_user.id
        )

        db.session.add(gasto)
        db.session.commit()

        return redirect(
            url_for("gastos")
        )

    return render_template(
        "nuevo.html"
    )

@app.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar_gasto(id):

    gasto = Gasto.query.get_or_404(id)

    if request.method == "POST":

        gasto.fecha = datetime.strptime(
            request.form["fecha"],
            "%Y-%m-%d"
        )

        gasto.pagado_a = request.form["pagado_a"]
        gasto.concepto = request.form["concepto"]
        gasto.observaciones = request.form["observaciones"]
        gasto.responsable = request.form["responsable"]
        gasto.importe = float(
            request.form["importe"]
        )

        db.session.commit()

        return redirect(
            url_for("gastos")
        )

    return render_template(
        "editar.html",
        gasto=gasto
    )
@app.route("/cambiar_comprobante/<int:id>")
@login_required
def cambiar_comprobante(id):

    gasto = Gasto.query.get_or_404(id)

    # Cambia el estado
    gasto.comprobante = not gasto.comprobante

    db.session.commit()

    return redirect(
        url_for(
            "gastos",
            fecha_inicio=request.args.get("fecha_inicio"),
            fecha_fin=request.args.get("fecha_fin"),
            responsable=request.args.get("responsable"),
            buscar=request.args.get("buscar"),
            orden=request.args.get("orden"),
            estado=request.args.get("estado")
        )
    )
@app.route("/eliminar/<int:id>")
@login_required
def eliminar_gasto(id):

    if current_user.rol != "Administrador":
        return "No tiene permisos para eliminar registros."

    gasto = Gasto.query.get_or_404(id)

    db.session.delete(gasto)
    db.session.commit()

    return redirect(url_for("gastos"))

@app.route("/exportar_excel")
@login_required
def exportar_excel():

    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    responsable = request.args.get("responsable")
    buscar = request.args.get("buscar")
    orden = request.args.get("orden", "fecha_desc")
    estado = request.args.get("estado", "todos")

    consulta = Gasto.query

    if fecha_inicio:
        consulta = consulta.filter(
            Gasto.fecha >= datetime.strptime(
                fecha_inicio,
                "%Y-%m-%d"
            ).date()
        )

    if fecha_fin:
        consulta = consulta.filter(
            Gasto.fecha <= datetime.strptime(
                fecha_fin,
                "%Y-%m-%d"
            ).date()
        )

    if responsable:
        consulta = consulta.filter(
            Gasto.responsable == responsable
        )

    if buscar:
        consulta = consulta.filter(
            or_(
                Gasto.pagado_a.contains(buscar),
                Gasto.concepto.contains(buscar),
                Gasto.observaciones.contains(buscar)
            )
        )

    # Filtrar por estado

    if estado == "pendientes":
        consulta = consulta.filter(
        Gasto.comprobante == False
    )

    elif estado == "elaborados":
        consulta = consulta.filter(
        Gasto.comprobante == True
    )   

    if orden == "fecha_asc":
        consulta = consulta.order_by(Gasto.fecha.asc())

    elif orden == "fecha_desc":
        consulta = consulta.order_by(Gasto.fecha.desc())

    elif orden == "importe_asc":
        consulta = consulta.order_by(Gasto.importe.asc())

    elif orden == "importe_desc":
        consulta = consulta.order_by(Gasto.importe.desc())

    elif orden == "responsable_asc":
        consulta = consulta.order_by(Gasto.responsable.asc())

    elif orden == "responsable_desc":
        consulta = consulta.order_by(Gasto.responsable.desc())

    else:
        consulta = consulta.order_by(Gasto.id.desc())

    gastos = consulta.all()


    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    encabezados = [
        "ID",
        "Fecha",
        "Pagado a",
        "Concepto",
        "Observaciones",
        "Responsable",
        "Importe",
        "Estado"
    ]

    ws.append(encabezados)

    for celda in ws[1]:
        celda.font = Font(bold=True)

    total = 0

    for g in gastos:

        ws.append([
            g.id,
            g.fecha.strftime("%d/%m/%Y"),
            g.pagado_a,
            g.concepto,
            g.observaciones,
            g.responsable,
            g.importe,
            "ELABORADO" if g.comprobante else "PENDIENTE"
        ])

        total += g.importe

    ws.append([])
    ws.append([
        "",
        "",
        "",
        "",
        "",
        "Total:",
        total
    ])

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    nombre = (
        f"gastos_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/importar_excel", methods=["GET", "POST"])
@login_required
def importar_excel():

    if request.method == "POST":

        archivo = request.files["archivo"]

        if archivo.filename == "":
            return "No seleccionó ningún archivo."

        wb = load_workbook(archivo)
        ws = wb.active

        registros = 0

        for fila in ws.iter_rows(min_row=2, values_only=True):

            print(fila)

            if not fila:
                continue

            try:

                fecha = fila[1]

                if fecha is None:
                    continue

                if isinstance(fecha, str):

                    formatos = [
                        "%d/%m/%Y",
                        "%Y-%m-%d"
                    ]

                    fecha_convertida = None

                    for formato in formatos:
                        try:
                            fecha_convertida = datetime.strptime(
                                fecha,
                                formato
                            ).date()
                            break
                        except ValueError:
                            pass

                    if fecha_convertida is None:
                        print(
                            f"No se pudo convertir la fecha: {fecha}"
                        )
                        continue

                    fecha = fecha_convertida

                elif isinstance(fecha, datetime):
                    fecha = fecha.date()

                gasto = Gasto(
                    fecha=fecha,
                    pagado_a=fila[2] or "",
                    concepto=fila[3] or "",
                    observaciones=fila[4] or "",
                    responsable=fila[5] or "",
                    importe=float(fila[6] or 0),
                    usuario_id=current_user.id
                )

                db.session.add(gasto)
                registros += 1

            except Exception as e:
                print(f"Error en fila: {fila}")
                print(e)

        db.session.commit()

        return f"Se importaron {registros} registros."

    return render_template("importar.html")

@app.route("/usuarios")
@login_required
def usuarios():

    if current_user.rol != "Administrador":
        return "No tiene permisos."

    lista = Usuario.query.order_by(
        Usuario.nombre
    ).all()

    return render_template(
        "usuarios.html",
        usuarios=lista
    )

@app.route("/usuarios/nuevo", methods=["GET", "POST"])
@login_required
def nuevo_usuario():

    if current_user.rol != "Administrador":
        return "No tiene permisos."

    if request.method == "POST":

        usuario = Usuario(
            usuario=request.form["usuario"],
            nombre=request.form["nombre"],
            rol=request.form["rol"],
            password=generate_password_hash(
                request.form["password"]
            )
        )

        db.session.add(usuario)
        db.session.commit()

        return redirect(
            url_for("usuarios")
        )

    return render_template(
        "nuevo_usuario.html"
    )

@app.route("/usuarios/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar_usuario(id):

    if current_user.rol != "Administrador":
        return "No tiene permisos."

    usuario = Usuario.query.get_or_404(id)

    if request.method == "POST":
        usuario.nombre = request.form["nombre"]
        usuario.usuario = request.form["usuario"]
        usuario.rol = request.form["rol"]

        if request.form["password"]:
            usuario.password = generate_password_hash(
                request.form["password"]
            )

        db.session.commit()

        return redirect(url_for("usuarios"))

    return render_template(
        "editar_usuario.html",
        usuario=usuario
    )

@app.route("/usuarios/eliminar/<int:id>")
@login_required
def eliminar_usuario(id):

    if current_user.rol != "Administrador":
        return "No tiene permisos."

    usuario = Usuario.query.get_or_404(id)

    if usuario.usuario == "admin":
        return "No se puede eliminar el administrador principal."

    db.session.delete(usuario)
    db.session.commit()

    return redirect(url_for("usuarios"))

if __name__ == "__main__":

    with app.app_context():

        db.create_all()

        if not Usuario.query.filter_by(usuario="admin").first():

            admin = Usuario(
                usuario="admin",
                nombre="Administrador",
                rol="Administrador",
                password=generate_password_hash("admin123")
            )

            db.session.add(admin)
            db.session.commit()

    app.run(debug=True)